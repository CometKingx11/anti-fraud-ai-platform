"""
批量导入用户功能
支持从 CSV 和 Excel (.xlsx) 文件批量导入用户
"""

import csv
import os
from app import create_app, db
from config.settings import config
from app.models.user import User
from app.utils.helpers import validate_student_id

# 尝试导入 openpyxl 用于处理 Excel 文件
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def import_users_from_file(file_path):
    """
    从文件批量导入用户（支持 CSV 和 Excel）
    
    Args:
        file_path: 文件路径（CSV 或 XLSX）
        
    Returns:
        dict: 导入结果统计
    """
    # 根据文件扩展名选择处理方式
    file_ext = os.path.splitext(file_path)[1].lower()
    
    if file_ext == '.xlsx':
        if not HAS_OPENPYXL:
            return {
                'success': 0,
                'failed': 0,
                'errors': ['缺少 openpyxl 库，无法处理 Excel 文件。请安装：pip install openpyxl']
            }
        return import_users_from_excel(file_path)
    else:
        # 默认使用 CSV 方式
        return import_users_from_csv(file_path)


def import_users_from_csv(file_path):
    """
    从 CSV 文件批量导入用户
    
    Args:
        file_path: CSV 文件路径
        
    Returns:
        dict: 导入结果统计
    """
    result = {
        'success': 0,
        'failed': 0,
        'errors': []
    }
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            
            for row in reader:
                try:
                    student_id = row.get('student_id', '').strip()
                    name = row.get('name', '').strip()
                    email = row.get('email', '').strip()
                    role = row.get('role', 'student').strip()
                    
                    # 处理密码：如果 CSV 中没有 password 列或该列为空，使用默认密码
                    password = row.get('password', '').strip()
                    if not password:
                        password = '123456'  # 默认密码
                    
                    # 验证学号
                    if not validate_student_id(student_id):
                        result['failed'] += 1
                        result['errors'].append(f"学号格式错误：{student_id}")
                        continue
                    
                    # 检查是否已存在
                    existing = User.get_by_student_id(student_id)
                    if existing:
                        result['failed'] += 1
                        result['errors'].append(f"学号已存在：{student_id}")
                        continue
                    
                    # 创建用户
                    User.create_user(
                        student_id=student_id,
                        password=password,
                        role=role,
                        name=name
                    )
                    result['success'] += 1
                    
                except Exception as e:
                    result['failed'] += 1
                    result['errors'].append(f"导入失败 {student_id}: {str(e)}")
        
        return result
        
    except Exception as e:
        result['errors'].append(f"读取文件失败：{str(e)}")
        return result


def import_users_from_excel(file_path):
    """
    从 Excel (.xlsx) 文件批量导入用户
    
    Args:
        file_path: Excel 文件路径
        
    Returns:
        dict: 导入结果统计
    """
    result = {
        'success': 0,
        'failed': 0,
        'errors': []
    }
    
    try:
        # 读取 Excel 文件
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active  # 使用第一个工作表
        
        # 读取表头
        headers = []
        for cell in ws[1]:
            headers.append(cell.value.strip() if cell.value else '')
        
        # 验证必要的列
        required_headers = ['student_id', 'name']
        for header in required_headers:
            if header not in headers:
                result['errors'].append(f"缺少必要的列：{header}")
                return result
        
        # 从第二行开始读取数据
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # 创建字典
                row_dict = dict(zip(headers, row))
                
                student_id = str(row_dict.get('student_id', '')).strip()
                name = str(row_dict.get('name', '')).strip() if row_dict.get('name') else ''
                email = str(row_dict.get('email', '')).strip() if row_dict.get('email') else ''
                role = str(row_dict.get('role', 'student')).strip() if row_dict.get('role') else 'student'
                
                # 处理密码
                password = str(row_dict.get('password', '')).strip() if row_dict.get('password') else ''
                if not password:
                    password = '123456'  # 默认密码
                
                # 验证学号
                if not validate_student_id(student_id):
                    result['failed'] += 1
                    result['errors'].append(f"学号格式错误 (第{row_idx}行): {student_id}")
                    continue
                
                # 检查是否已存在
                existing = User.get_by_student_id(student_id)
                if existing:
                    result['failed'] += 1
                    result['errors'].append(f"学号已存在 (第{row_idx}行): {student_id}")
                    continue
                
                # 创建用户
                User.create_user(
                    student_id=student_id,
                    password=password,
                    role=role,
                    name=name
                )
                result['success'] += 1
                
            except Exception as e:
                result['failed'] += 1
                result['errors'].append(f"导入失败 (第{row_idx}行): {str(e)}")
        
        wb.close()
        return result
        
    except Exception as e:
        result['errors'].append(f"读取 Excel 文件失败：{str(e)}")
        return result


if __name__ == '__main__':
    app = create_app(config[os.environ.get('FLASK_ENV', 'default')])
    
    with app.app_context():
        print("=" * 60)
        print("批量导入用户测试")
        print("=" * 60)
        
        # 示例文件路径（支持 CSV 和 XLSX）
        test_files = ['users_import.csv', 'users_import.xlsx']
        
        for import_file in test_files:
            if os.path.exists(import_file):
                print(f"\n正在从 {import_file} 导入用户...")
                result = import_users_from_file(import_file)
                
                print(f"\n导入完成！")
                print(f"成功：{result['success']} 人")
                print(f"失败：{result['failed']} 人")
                
                if result['errors']:
                    print("\n错误详情:")
                    for error in result['errors']:
                        print(f"  - {error}")
                
                print("\n" + "=" * 60)
                break
        else:
            print(f"\n未找到导入文件：{', '.join(test_files)}")
            print("\n请创建 CSV 或 XLSX 文件，格式如下:")
            print("student_id,name,email,role,password")
            print("20240001，张三，zhangsan@example.com,student,123456")
            print("20240002，李四，lisi@example.com,student,123456")
        
        print("\n" + "=" * 60)
