"""
批量导入问卷题目服务
支持从 CSV 和 Excel (.xlsx) 文件批量导入问卷题目
"""

import csv
import os
import json
from app import create_app, db
from config.settings import config
from app.models.questionnaire import QuestionnaireQuestion

# 尝试导入 openpyxl 用于处理 Excel 文件
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def import_questions_from_file(file_path):
    """
    从文件批量导入题目（支持 CSV 和 Excel）
    
    Args:
        file_path: 文件路径（CSV 或 XLSX）
        
    Returns:
        dict: 导入结果统计
    """
    # 根据文件扩展名选择处理方式
    file_ext = os.path.splitext(file_path)[1].lower()
    
    if file_ext == '.xlsx':
        if not HAS_OPENPYXL:
            return {
                'success': 0,
                'failed': 0,
                'errors': ['缺少 openpyxl 库，无法处理 Excel 文件。请安装：pip install openpyxl']
            }
        return import_questions_from_excel(file_path)
    else:
        # 默认使用 CSV 方式
        return import_questions_from_csv(file_path)


def import_questions_from_csv(file_path):
    """
    从 CSV 文件批量导入题目
    
    CSV 格式示例：
    question_number,category,question_text,min_score,max_score,dimension,weight,is_active,is_required,display_order,options_json
    1,认知维度，你是否了解什么是电信诈骗，0,10,cognitive,1.0,True,True,1,"{""A"": 0, ""B"": 10}"
    
    Args:
        file_path: CSV 文件路径
        
    Returns:
        dict: 导入结果统计
    """
    result = {
        'success': 0,
        'failed': 0,
        'errors': []
    }
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            
            for row in reader:
                try:
                    # 验证必要的列
                    required_fields = ['question_number', 'category', 'question_text', 'min_score', 'max_score', 'dimension']
                    missing_fields = [field for field in required_fields if field not in row or not row[field]]
                    
                    if missing_fields:
                        result['failed'] += 1
                        result['errors'].append(f"缺少必要字段：{', '.join(missing_fields)} (第{reader.line_num}行)")
                        continue
                    
                    # 检查题目编号是否已存在
                    existing = QuestionnaireQuestion.query.filter_by(
                        question_number=int(row['question_number']),
                        category=row['category']
                    ).first()
                    
                    if existing:
                        result['failed'] += 1
                        result['errors'].append(f"题目已存在：{row['question_number']} - {row['category']} (第{reader.line_num}行)")
                        continue
                    
                    # 创建题目
                    question = QuestionnaireQuestion(
                        question_number=int(row['question_number']),
                        category=row['category'],
                        question_text=row['question_text'],
                        min_score=int(row['min_score']),
                        max_score=int(row['max_score']),
                        dimension=row['dimension'],
                        weight=float(row.get('weight', 1.0) or 1.0),
                        options_json=row.get('options_json', '{}'),
                        is_active=row.get('is_active', 'True').lower() in ['true', '1', 'yes'],
                        is_required=row.get('is_required', 'True').lower() in ['true', '1', 'yes'],
                        display_order=int(row.get('display_order', 0) or 0)
                    )
                    
                    db.session.add(question)
                    result['success'] += 1
                    
                except Exception as e:
                    result['failed'] += 1
                    result['errors'].append(f"导入失败 (第{reader.line_num}行): {str(e)}")
        
        # 提交所有题目
        db.session.commit()
        return result
        
    except Exception as e:
        db.session.rollback()
        result['errors'].append(f"读取文件失败：{str(e)}")
        return result


def import_questions_from_excel(file_path):
    """
    从 Excel (.xlsx) 文件批量导入题目
    
    Args:
        file_path: Excel 文件路径
        
    Returns:
        dict: 导入结果统计
    """
    result = {
        'success': 0,
        'failed': 0,
        'errors': []
    }
    
    try:
        # 读取 Excel 文件
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active  # 使用第一个工作表
        
        # 读取表头
        headers = []
        for cell in ws[1]:
            headers.append(cell.value.strip() if cell.value else '')
        
        # 验证必要的列
        required_headers = ['question_number', 'category', 'question_text', 'min_score', 'max_score', 'dimension']
        for header in required_headers:
            if header not in headers:
                result['errors'].append(f"缺少必要的列：{header}")
                return result
        
        # 从第二行开始读取数据
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # 创建字典
                row_dict = dict(zip(headers, row))
                
                # 验证必要的字段
                missing_fields = [field for field in required_headers if field not in row_dict or not row_dict[field]]
                if missing_fields:
                    result['failed'] += 1
                    result['errors'].append(f"缺少必要字段：{', '.join(missing_fields)} (第{row_idx}行)")
                    continue
                
                # 检查题目编号是否已存在
                existing = QuestionnaireQuestion.query.filter_by(
                    question_number=int(row_dict['question_number']),
                    category=row_dict['category']
                ).first()
                
                if existing:
                    result['failed'] += 1
                    result['errors'].append(f"题目已存在：{row_dict['question_number']} - {row_dict['category']} (第{row_idx}行)")
                    continue
                
                # 创建题目
                question = QuestionnaireQuestion(
                    question_number=int(row_dict['question_number']),
                    category=row_dict['category'],
                    question_text=row_dict['question_text'],
                    min_score=int(row_dict['min_score']),
                    max_score=int(row_dict['max_score']),
                    dimension=row_dict['dimension'],
                    weight=float(row_dict.get('weight', 1.0) or 1.0),
                    options_json=row_dict.get('options_json', '{}') or '{}',
                    is_active=str(row_dict.get('is_active', 'True')).lower() in ['true', '1', 'yes'],
                    is_required=str(row_dict.get('is_required', 'True')).lower() in ['true', '1', 'yes'],
                    display_order=int(row_dict.get('display_order', 0) or 0)
                )
                
                db.session.add(question)
                result['success'] += 1
                
            except Exception as e:
                result['failed'] += 1
                result['errors'].append(f"导入失败 (第{row_idx}行): {str(e)}")
        
        # 提交所有题目
        db.session.commit()
        wb.close()
        return result
        
    except Exception as e:
        db.session.rollback()
        result['errors'].append(f"读取 Excel 文件失败：{str(e)}")
        return result


if __name__ == '__main__':
    app = create_app(config[os.environ.get('FLASK_ENV', 'default')])
    
    with app.app_context():
        print("=" * 60)
        print("批量导入问卷题目测试")
        print("=" * 60)
        
        # 示例文件路径（支持 CSV 和 XLSX）
        test_files = ['questions_import.csv', 'questions_import.xlsx']
        
        for import_file in test_files:
            if os.path.exists(import_file):
                print(f"\n正在从 {import_file} 导入题目...")
                result = import_questions_from_file(import_file)
                
                print(f"\n导入完成！")
                print(f"成功：{result['success']} 题")
                print(f"失败：{result['failed']} 题")
                
                if result['errors']:
                    print("\n错误详情:")
                    for error in result['errors']:
                        print(f"  - {error}")
                
                print("\n" + "=" * 60)
                break
        else:
            print(f"\n未找到导入文件：{', '.join(test_files)}")
            print("\n请创建 CSV 文件，格式如下:")
            print("question_number,category,question_text,min_score,max_score,dimension,weight,is_active,is_required,display_order,options_json")
            print("1，认知维度，你是否了解什么是电信诈骗，0,10,cognitive,1.0,True,True,1,\"{""A"": 0, ""B"": 10}\"")
        
        print("\n" + "=" * 60)
